"""Rendering puro (senza FastAPI) del rendiconto in PDF e XLSX.

Le due funzioni ``render_rendiconto_pdf`` / ``render_rendiconto_xlsx``
ricevono un :class:`RendicontoResponse` già aggregato e restituiscono i byte
del file, così la logica di rendering resta isolata e facilmente testabile.
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    KeepTogether,
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from app.schemas.rendiconto import RendicontoResponse, SezioneRendicontoAggregato

EXCEL_EURO_FORMAT = "€ #,##0.00"

_GREEN = "FF2E7D32"
_RED = "FFC62828"


def format_euro(d: Decimal) -> str:
    """Formatta un importo in stile italiano: ``€ 1.234,56``.

    Punto come separatore delle migliaia, virgola per i decimali.
    """
    q = Decimal(d).quantize(Decimal("0.01"))
    # Formatta in stile US (virgola migliaia, punto decimale) poi inverte i
    # separatori usando un segnaposto temporaneo.
    s = f"{q:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")
    return f"€ {s}"


def _is_fuori_bilancio(sezione: SezioneRendicontoAggregato) -> bool:
    return sezione.descrizione.strip().lower() == "fuori bilancio"


def _has_activity(sezione: SezioneRendicontoAggregato) -> bool:
    """True se la sezione ha movimenti (totale o sottovoci non a zero)."""
    if sezione.totale != 0:
        return True
    for voce in sezione.voci:
        if voce.totale != 0:
            return True
        for sottovoce in voce.sottovoci:
            if sottovoce.totale != 0:
                return True
    return False


def _visible_sezioni(
    data: RendicontoResponse,
) -> list[SezioneRendicontoAggregato]:
    """Sezioni da mostrare: salta le vuote, ma tiene sempre Fuori Bilancio."""
    return [s for s in data.sezioni if _has_activity(s) or _is_fuori_bilancio(s)]


# ── PDF ───────────────────────────────────────────────────────────────────────


def render_rendiconto_pdf(data: RendicontoResponse) -> bytes:
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        topMargin=18 * mm,
        bottomMargin=18 * mm,
        title=f"Rendiconto Anno {data.anno}",
    )

    base = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "RendTitle", parent=base["Title"], fontSize=16, alignment=TA_CENTER
    )
    subtitle_style = ParagraphStyle(
        "RendSubtitle",
        parent=base["Normal"],
        fontSize=11,
        alignment=TA_CENTER,
        spaceAfter=2,
    )
    sez_label = ParagraphStyle(
        "SezLabel", parent=base["Normal"], fontSize=12, fontName="Helvetica-Bold"
    )
    sez_amount = ParagraphStyle(
        "SezAmount",
        parent=base["Normal"],
        fontSize=12,
        fontName="Helvetica-Bold",
        alignment=TA_RIGHT,
    )
    voce_label = ParagraphStyle(
        "VoceLabel",
        parent=base["Normal"],
        fontSize=10,
        fontName="Helvetica-Bold",
        leftIndent=10,
    )
    voce_amount = ParagraphStyle(
        "VoceAmount", parent=base["Normal"], fontSize=10, alignment=TA_RIGHT
    )
    sv_label = ParagraphStyle(
        "SvLabel", parent=base["Normal"], fontSize=9, leftIndent=22
    )
    sv_amount = ParagraphStyle(
        "SvAmount", parent=base["Normal"], fontSize=9, alignment=TA_RIGHT
    )

    usable = doc.width
    label_w = usable * 0.74
    amount_w = usable - label_w

    elements: list = []
    elements.append(Paragraph(f"Rendiconto Anno {data.anno}", title_style))
    elements.append(Paragraph(data.banda_nome, subtitle_style))
    elements.append(Spacer(1, 8))

    # ── Saldi iniziali ────────────────────────────────────────────────────
    saldi_data = [
        ["Saldo iniziale cassa", format_euro(data.saldo_iniziale_cassa)],
        ["Saldo iniziale banca", format_euro(data.saldo_iniziale_banca)],
    ]
    saldi_tbl = Table(saldi_data, colWidths=[label_w, amount_w])
    saldi_tbl.setStyle(
        TableStyle(
            [
                ("FONTSIZE", (0, 0), (-1, -1), 10),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.lightgrey),
                ("TOPPADDING", (0, 0), (-1, -1), 3),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
            ]
        )
    )
    elements.append(saldi_tbl)
    elements.append(Spacer(1, 12))

    # ── Sezioni ───────────────────────────────────────────────────────────
    for sezione in _visible_sezioni(data):
        rows = [
            [
                Paragraph(sezione.descrizione, sez_label),
                Paragraph(f"Totale: {format_euro(sezione.totale)}", sez_amount),
            ]
        ]
        for voce in sezione.voci:
            rows.append(
                [
                    Paragraph(voce.descrizione, voce_label),
                    Paragraph(format_euro(voce.totale), voce_amount),
                ]
            )
            for sottovoce in voce.sottovoci:
                rows.append(
                    [
                        Paragraph(sottovoce.descrizione, sv_label),
                        Paragraph(format_euro(sottovoce.totale), sv_amount),
                    ]
                )
        sez_tbl = Table(rows, colWidths=[label_w, amount_w])
        sez_tbl.setStyle(
            TableStyle(
                [
                    ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.grey),
                    ("TOPPADDING", (0, 0), (-1, -1), 2),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]
            )
        )

        sez_elements = [sez_tbl, Spacer(1, 10)]

        if "entrate" in sezione.descrizione.lower():
            elements.append(KeepTogether(sez_elements))
        else:
            elements.extend(sez_elements)

    # ── Totali generali ───────────────────────────────────────────────────
    elements.append(Spacer(1, 6))
    avanzo = data.totali.avanzo_disavanzo
    saldo_finale = data.totali.saldo_finale_cassa + data.totali.saldo_finale_banca
    totali_rows = [
        ["Totali generali", ""],
        ["Totale entrate", format_euro(data.totali.entrate)],
        ["Totale uscite", format_euro(data.totali.uscite)],
        ["Avanzo/Disavanzo", format_euro(avanzo)],
        ["Saldo finale (cassa + banca)", format_euro(saldo_finale)],
    ]
    totali_tbl = Table(totali_rows, colWidths=[label_w, amount_w])
    totali_style = [
        ("FONTSIZE", (0, 0), (-1, -1), 10),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("SPAN", (0, 0), (1, 0)),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.grey),
        ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.grey),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]
    if avanzo > 0:
        totali_style.append(("TEXTCOLOR", (1, 3), (1, 3), colors.green))
    elif avanzo < 0:
        totali_style.append(("TEXTCOLOR", (1, 3), (1, 3), colors.red))
    totali_tbl.setStyle(TableStyle(totali_style))
    elements.append(totali_tbl)

    def _footer(canvas, doc_) -> None:
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        generato = datetime.now().strftime("%d/%m/%Y %H:%M")
        canvas.drawCentredString(A4[0] / 2, 10 * mm, f"Generato il {generato}")
        canvas.restoreState()

    doc.build(elements, onFirstPage=_footer, onLaterPages=_footer)
    return buffer.getvalue()


# ── XLSX ──────────────────────────────────────────────────────────────────────


def render_rendiconto_xlsx(data: RendicontoResponse) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Rendiconto {data.anno}"

    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 20

    bold = Font(bold=True)
    title_font = Font(bold=True, size=16)
    center = Alignment(horizontal="center")

    # Row 1 — titolo
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"Rendiconto Anno {data.anno}"
    c.font = title_font
    c.alignment = center

    # Row 2 — banda
    ws["A2"] = data.banda_nome

    # Row 3 vuota → saldi iniziali sulle righe 4-5
    def _euro_cell(row: int, col: int, value: Decimal) -> None:
        cell = ws.cell(row=row, column=col, value=float(value))
        cell.number_format = EXCEL_EURO_FORMAT

    ws.cell(row=4, column=1, value="Saldo iniziale cassa")
    _euro_cell(4, 2, data.saldo_iniziale_cassa)
    ws.cell(row=5, column=1, value="Saldo iniziale banca")
    _euro_cell(5, 2, data.saldo_iniziale_banca)

    # Row 6 vuota → intestazione tabella sezioni su riga 7
    row = 7
    header_label = ws.cell(row=row, column=1, value="Sezione / Voce / Sottovoce")
    header_label.font = bold
    header_amount = ws.cell(row=row, column=2, value="Importo")
    header_amount.font = bold
    row += 1

    for sezione in _visible_sezioni(data):
        sez_cell = ws.cell(row=row, column=1, value=sezione.descrizione)
        sez_cell.font = bold
        sez_amount = ws.cell(row=row, column=2, value=float(sezione.totale))
        sez_amount.number_format = EXCEL_EURO_FORMAT
        sez_amount.font = bold
        row += 1
        for voce in sezione.voci:
            ws.cell(row=row, column=1, value=f"  {voce.descrizione}")
            _euro_cell(row, 2, voce.totale)
            row += 1
            for sottovoce in voce.sottovoci:
                ws.cell(row=row, column=1, value=f"    {sottovoce.descrizione}")
                _euro_cell(row, 2, sottovoce.totale)
                row += 1

    # Riga vuota poi totali
    row += 1
    avanzo = data.totali.avanzo_disavanzo
    saldo_finale = data.totali.saldo_finale_cassa + data.totali.saldo_finale_banca

    def _totali_row(label: str, value: Decimal) -> int:
        nonlocal row
        lbl = ws.cell(row=row, column=1, value=label)
        lbl.font = bold
        val = ws.cell(row=row, column=2, value=float(value))
        val.number_format = EXCEL_EURO_FORMAT
        current = row
        row += 1
        return current

    _totali_row("Totale entrate", data.totali.entrate)
    _totali_row("Totale uscite", data.totali.uscite)
    avanzo_row = _totali_row("Avanzo/Disavanzo", avanzo)
    _totali_row("Saldo finale (cassa + banca)", saldo_finale)

    if avanzo > 0:
        ws.cell(row=avanzo_row, column=2).fill = PatternFill(
            start_color=_GREEN, end_color=_GREEN, fill_type="solid"
        )
    elif avanzo < 0:
        ws.cell(row=avanzo_row, column=2).fill = PatternFill(
            start_color=_RED, end_color=_RED, fill_type="solid"
        )

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
