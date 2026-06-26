from __future__ import annotations

from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from httpx import AsyncClient

from app.utils.export_rendiconto import format_euro

CFG_BASE = "/api/v1/configurazioni-banda-anno"
FLUSSO_BASE = "/api/v1/flussi-cassa"
PDF_URL = "/api/v1/contabilita/rendiconto/export/pdf"
XLSX_URL = "/api/v1/contabilita/rendiconto/export/xlsx"

PDF_MEDIA_TYPE = "application/pdf"
XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def setup_env_with_flussi(
    client: AsyncClient, banda_codice: int = 1, anno: int = 2026
) -> dict:
    """Seed nature + config (seeds voci) and add a couple of flussi so the
    export has real content."""
    await client.post(
        "/api/v1/nature-flusso/", json={"codice": 1, "descrizione": "Banca"}
    )
    await client.post(
        "/api/v1/nature-flusso/", json={"codice": 2, "descrizione": "Cassa"}
    )
    cfg_resp = await client.post(
        f"{CFG_BASE}/", json={"banda_codice": banda_codice, "anno": anno}
    )
    assert cfg_resp.status_code == 201, cfg_resp.text
    cfg = cfg_resp.json()

    voci_resp = await client.get(
        f"/api/v1/voci-contabilita/?banda_codice={banda_codice}&limit=100"
    )
    voci = {v["voce_contabilita"]: v for v in voci_resp.json()["items"]}

    # Entrata 100 + uscita 30
    await client.post(
        f"{FLUSSO_BASE}/",
        json={
            "data_registrazione": f"{anno}-02-10T00:00:00",
            "descrizione_operazione": "Quota",
            "voce_contabilita_id": voci["Quote associative"]["id"],
            "importo": 100.0,
            "segno": "+",
            "natura_flusso_codice": 1,
        },
    )
    await client.post(
        f"{FLUSSO_BASE}/",
        json={
            "data_registrazione": f"{anno}-03-10T00:00:00",
            "descrizione_operazione": "Spesa",
            "voce_contabilita_id": voci["Spese bancarie"]["id"],
            "importo": 30.0,
            "segno": "-",
            "natura_flusso_codice": 1,
        },
    )
    return {"cfg": cfg, "voci": voci}


# ── format_euro (pure) ─────────────────────────────────────────────────────


def test_format_euro_italian_locale():
    assert format_euro(Decimal("1234.56")) == "€ 1.234,56"
    assert format_euro(Decimal("0")) == "€ 0,00"
    assert format_euro(Decimal("-1234.56")) == "€ -1.234,56"
    assert format_euro(Decimal("1000000")) == "€ 1.000.000,00"


# ── PDF ─────────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_export_pdf_returns_pdf_bytes(client: AsyncClient):
    await setup_env_with_flussi(client)
    resp = await client.get(f"{PDF_URL}?banda_codice=1&anno=2026")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == PDF_MEDIA_TYPE
    assert resp.content[:4] == b"%PDF"


@pytest.mark.asyncio
async def test_export_pdf_filename_header_correct(client: AsyncClient):
    await setup_env_with_flussi(client)
    resp = await client.get(f"{PDF_URL}?banda_codice=1&anno=2026")
    assert resp.status_code == 200
    assert (
        resp.headers["content-disposition"]
        == 'attachment; filename="rendiconto_1_2026.pdf"'
    )


@pytest.mark.asyncio
async def test_export_pdf_size_reasonable(client: AsyncClient):
    await setup_env_with_flussi(client)
    resp = await client.get(f"{PDF_URL}?banda_codice=1&anno=2026")
    assert resp.status_code == 200
    assert len(resp.content) > 1000


@pytest.mark.asyncio
async def test_export_pdf_no_data_doesnt_crash(client: AsyncClient):
    # No config, no flussi → rendiconto is empty but PDF must still render.
    resp = await client.get(f"{PDF_URL}?banda_codice=99&anno=2099")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == PDF_MEDIA_TYPE
    assert resp.content[:4] == b"%PDF"
    assert len(resp.content) > 1000


# ── XLSX ────────────────────────────────────────────────────────────────────


@pytest.mark.asyncio
async def test_export_xlsx_returns_xlsx_bytes(client: AsyncClient):
    await setup_env_with_flussi(client)
    resp = await client.get(f"{XLSX_URL}?banda_codice=1&anno=2026")
    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_MEDIA_TYPE
    # XLSX is a zip archive → PK\x03\x04 magic.
    assert resp.content[:4] == b"PK\x03\x04"


@pytest.mark.asyncio
async def test_export_xlsx_filename_header_correct(client: AsyncClient):
    await setup_env_with_flussi(client)
    resp = await client.get(f"{XLSX_URL}?banda_codice=1&anno=2026")
    assert resp.status_code == 200
    assert (
        resp.headers["content-disposition"]
        == 'attachment; filename="rendiconto_1_2026.xlsx"'
    )


@pytest.mark.asyncio
async def test_export_xlsx_openable_by_openpyxl(client: AsyncClient):
    await setup_env_with_flussi(client)
    resp = await client.get(f"{XLSX_URL}?banda_codice=1&anno=2026")
    assert resp.status_code == 200

    wb = openpyxl.load_workbook(BytesIO(resp.content))
    assert "Rendiconto 2026" in wb.sheetnames
    ws = wb["Rendiconto 2026"]
    assert ws["A1"].value == "Rendiconto Anno 2026"
